"""
Bulk import / template download endpoints for Products.

Admin uploads an .xlsx or .csv file; rows are parsed, validated, and
turned into Product rows. Categories and Manufacturers are matched by
slug (preferred) or name (fallback).

Also exposes a /template/ endpoint that streams back a ready-to-fill
.xlsx template with column headers + one example row.
"""
from __future__ import annotations

import csv
import io
import uuid
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.http import HttpResponse
from django.utils.text import slugify
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView

from products.models import Category, Manufacturer, Product, ProductImage


# Column definitions — keep in sync with frontend docs and the template.
COLUMNS = [
    ("name", "Tên sản phẩm", True),
    ("sku", "Mã SKU", True),
    ("price", "Giá gốc (VND)", True),
    ("sale_price", "Giá khuyến mãi (VND)", False),
    ("category", "Danh mục (slug hoặc tên)", True),
    ("manufacturer", "Nhà sản xuất (slug hoặc tên)", True),
    ("stock_quantity", "Tồn kho", False),
    ("unit", "Đơn vị (Hộp/Vỉ/Chai...)", False),
    ("quantity_per_unit", "Quy cách đóng gói", False),
    ("short_description", "Mô tả ngắn", False),
    ("description", "Mô tả chi tiết", False),
    ("ingredients", "Thành phần", False),
    ("dosage", "Liều dùng", False),
    ("usage", "Cách dùng", False),
    ("contraindications", "Chống chỉ định", False),
    ("side_effects", "Tác dụng phụ", False),
    ("storage", "Bảo quản", False),
    ("requires_prescription", "Cần đơn thuốc (true/false)", False),
    ("is_featured", "Nổi bật (true/false)", False),
    ("status", "Trạng thái (DRAFT/ACTIVE)", False),
    ("image_urls", "Link ảnh (phân cách bằng |)", False),
    ("meta_title", "SEO title", False),
    ("meta_description", "SEO description", False),
]

REQUIRED_KEYS = [key for key, _label, required in COLUMNS if required]
ALL_KEYS = [key for key, _label, _ in COLUMNS]


def _parse_bool(value) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ("true", "1", "yes", "y", "có", "co", "x")


def _parse_int(value, default=0) -> int:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _parse_decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError, ValueError):
        return None


def _read_rows_from_xlsx(file_bytes: bytes) -> list[dict]:
    """Parse an uploaded .xlsx into a list of dicts keyed by header."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []

    results: list[dict] = []
    for row in rows_iter:
        if row is None:
            continue
        # Skip completely empty rows
        if all(c is None or c == "" for c in row):
            continue
        record = {}
        for i, key in enumerate(header):
            if not key:
                continue
            value = row[i] if i < len(row) else None
            record[key] = value
        results.append(record)
    return results


def _read_rows_from_csv(file_bytes: bytes) -> list[dict]:
    """Parse a CSV upload (UTF-8 or UTF-8 BOM)."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    results: list[dict] = []
    for row in reader:
        if not any((v or "").strip() for v in row.values()):
            continue
        results.append({(k or "").strip(): v for k, v in row.items()})
    return results


def _resolve_category(value):
    if not value:
        return None
    value = str(value).strip()
    return (
        Category.objects.filter(slug=value).first()
        or Category.objects.filter(name__iexact=value).first()
    )


def _resolve_manufacturer(value):
    if not value:
        return None
    value = str(value).strip()
    return (
        Manufacturer.objects.filter(slug=value).first()
        or Manufacturer.objects.filter(name__iexact=value).first()
    )


def _build_product_from_row(row: dict, user) -> tuple[Product | None, list[str]]:
    """Return (product, errors). Product is unsaved if errors is empty."""
    errors: list[str] = []
    clean = {k: (row.get(k) if row.get(k) is not None else "") for k in ALL_KEYS}

    # Required fields
    for key in REQUIRED_KEYS:
        if clean.get(key) in (None, ""):
            errors.append(f"Thiếu trường bắt buộc: {key}")

    name = str(clean.get("name") or "").strip()
    sku = str(clean.get("sku") or "").strip()
    price = _parse_decimal(clean.get("price"))
    if clean.get("price") not in (None, "") and price is None:
        errors.append("Giá gốc không hợp lệ")

    sale_price = _parse_decimal(clean.get("sale_price"))

    category = _resolve_category(clean.get("category"))
    if clean.get("category") and category is None:
        errors.append(f"Không tìm thấy danh mục '{clean.get('category')}'")

    manufacturer = _resolve_manufacturer(clean.get("manufacturer"))
    if clean.get("manufacturer") and manufacturer is None:
        errors.append(f"Không tìm thấy nhà sản xuất '{clean.get('manufacturer')}'")

    if sku and Product.objects.filter(sku=sku).exists():
        errors.append(f"SKU '{sku}' đã tồn tại")

    if errors:
        return None, errors

    status_value = str(clean.get("status") or "").strip().upper() or Product.Status.DRAFT
    if status_value not in {s.value for s in Product.Status}:
        status_value = Product.Status.DRAFT

    product = Product(
        sku=sku,
        name=name,
        slug=slugify(name, allow_unicode=False) or str(uuid.uuid4())[:8],
        short_description=str(clean.get("short_description") or "")[:500],
        description=str(clean.get("description") or ""),
        price=price,
        sale_price=sale_price,
        category=category,
        manufacturer=manufacturer,
        ingredients=str(clean.get("ingredients") or ""),
        dosage=str(clean.get("dosage") or ""),
        usage=str(clean.get("usage") or ""),
        contraindications=str(clean.get("contraindications") or ""),
        side_effects=str(clean.get("side_effects") or ""),
        storage=str(clean.get("storage") or ""),
        unit=str(clean.get("unit") or "Hộp")[:50],
        quantity_per_unit=str(clean.get("quantity_per_unit") or "")[:100],
        stock_quantity=_parse_int(clean.get("stock_quantity")),
        status=status_value,
        requires_prescription=_parse_bool(clean.get("requires_prescription")),
        is_featured=_parse_bool(clean.get("is_featured")),
        meta_title=str(clean.get("meta_title") or "")[:200],
        meta_description=str(clean.get("meta_description") or "")[:500],
        created_by=user if user and user.is_authenticated else None,
    )
    return product, []


def _save_product_with_images(product: Product, row: dict):
    product.save()
    image_urls_raw = row.get("image_urls") or ""
    if image_urls_raw:
        urls = [u.strip() for u in str(image_urls_raw).split("|") if u.strip()]
        for i, url in enumerate(urls):
            ProductImage.objects.create(
                product=product,
                image_url=url,
                is_primary=(i == 0),
                sort_order=i,
                alt_text=product.name[:200],
            )


class ProductBulkImportView(APIView):
    """
    POST /api/admin/products/bulk-import/
    Body: multipart/form-data with 'file' = .xlsx | .csv
    Optional: 'dry_run' (true/false) — validate only, don't save

    Response:
      {
        "total_rows": int,
        "success_count": int,
        "error_count": int,
        "created": [ { "row": int, "sku": str, "name": str, "id": str } ],
        "errors": [ { "row": int, "messages": [str, ...] } ]
      }
    """

    permission_classes = [IsAdminUser]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        upload = request.FILES.get("file")
        if upload is None:
            return Response(
                {"detail": "Thiếu file upload (field 'file')."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        dry_run = str(request.data.get("dry_run", "")).lower() in ("1", "true", "yes")

        filename = (upload.name or "").lower()
        file_bytes = upload.read()
        try:
            if filename.endswith(".csv"):
                rows = _read_rows_from_csv(file_bytes)
            elif filename.endswith(".xlsx") or filename.endswith(".xlsm"):
                rows = _read_rows_from_xlsx(file_bytes)
            else:
                return Response(
                    {"detail": "Chỉ hỗ trợ file .xlsx hoặc .csv"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:  # noqa: BLE001
            return Response(
                {"detail": f"Không đọc được file: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created: list[dict] = []
        errors: list[dict] = []

        # Row numbering starts at 2 (row 1 = header)
        with transaction.atomic():
            savepoint = transaction.savepoint()
            for idx, row in enumerate(rows, start=2):
                product, row_errors = _build_product_from_row(row, request.user)
                if row_errors:
                    errors.append({"row": idx, "messages": row_errors})
                    continue
                if dry_run:
                    created.append(
                        {
                            "row": idx,
                            "sku": product.sku,
                            "name": product.name,
                            "id": None,
                        }
                    )
                    continue
                try:
                    _save_product_with_images(product, row)
                    created.append(
                        {
                            "row": idx,
                            "sku": product.sku,
                            "name": product.name,
                            "id": str(product.id),
                        }
                    )
                except Exception as e:  # noqa: BLE001
                    errors.append({"row": idx, "messages": [f"Lỗi khi lưu: {e}"]})

            if dry_run:
                transaction.savepoint_rollback(savepoint)
            else:
                transaction.savepoint_commit(savepoint)

        return Response(
            {
                "total_rows": len(rows),
                "success_count": len(created),
                "error_count": len(errors),
                "dry_run": dry_run,
                "created": created,
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )


class ProductBulkImportTemplateView(APIView):
    """
    GET /api/admin/products/bulk-import/template/
    Returns an .xlsx template with headers and one example row.
    """

    permission_classes = [IsAdminUser]

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment

        wb = Workbook()
        ws = wb.active
        ws.title = "Sản phẩm"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="16A34A")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header row
        for col_idx, (key, label, required) in enumerate(COLUMNS, start=1):
            header = f"{label}{' *' if required else ''}"
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            # Column width heuristic
            ws.column_dimensions[cell.column_letter].width = max(18, len(header) + 4)

        # Example row
        example = {
            "name": "Paralmax Extra Boston (hộp 180 viên)",
            "sku": "PARA-EXTRA-180",
            "price": 95000,
            "sale_price": 85000,
            "category": "thuoc-giam-dau",
            "manufacturer": "boston-pharma",
            "stock_quantity": 100,
            "unit": "Hộp",
            "quantity_per_unit": "15 vỉ x 12 viên",
            "short_description": "Thuốc giảm đau hạ sốt Paracetamol + Cafein.",
            "description": "Paralmax Extra là thuốc giảm đau - hạ sốt...\n- Đau đầu\n- Đau bụng kinh",
            "ingredients": "Paracetamol 500mg, Cafein 65mg",
            "dosage": "Người lớn: 1-2 viên x 3 lần/ngày",
            "usage": "Uống sau ăn",
            "contraindications": "Bệnh gan nặng, dị ứng Paracetamol",
            "side_effects": "Buồn nôn nhẹ, mẩn ngứa (hiếm)",
            "storage": "Nơi khô ráo, dưới 30°C",
            "requires_prescription": "false",
            "is_featured": "true",
            "status": "ACTIVE",
            "image_urls": "https://minio.banthuocsi.vn/banthuoc-storage/products/para1.jpg|https://minio.banthuocsi.vn/banthuoc-storage/products/para2.jpg",
            "meta_title": "Paralmax Extra Boston - Giảm đau hạ sốt hiệu quả",
            "meta_description": "Mua Paralmax Extra chính hãng Boston. Giao nhanh toàn quốc.",
        }
        for col_idx, (key, _label, _) in enumerate(COLUMNS, start=1):
            ws.cell(row=2, column=col_idx, value=example.get(key, ""))

        # Add a comment on the header explaining required marker
        ws["A1"].comment = Comment(
            "Các cột có dấu * là bắt buộc. Danh mục và Nhà sản xuất có thể nhập slug hoặc tên.",
            "BanThuoc",
        )

        # Freeze header
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="banthuoc-products-template.xlsx"'
        )
        return response
